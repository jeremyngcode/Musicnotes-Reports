from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from settings import *
# -------------------------------------------------------------------------------------------------

# Load Master Excel workbook
xl_wb = load_workbook(master_xl_file, read_only=True)
sheet = xl_wb.worksheets[0]

# Retrieve sheetmusic titles
master_list = []

for row in sheet.iter_rows(min_row=4, min_col=1, max_col=1, values_only=True):
	if row[0] is not None:
		master_list.append(row[0])
	else:
		break

# Load Musicnotes Excel workbook
xl_wb = load_workbook(musicnotes_xl_file, read_only=True)
sheet = xl_wb.worksheets[0]

# Retrieve sheetmusic revenue data
print('Retrieving data..')

musicnotes_data = {}

for row in sheet.iter_rows(min_row=5, min_col=2, max_col=6, values_only=True):
	if row[0] is not None:
		musicnotes_data[row[0].lower()] = {
			'Downloads': row[2],
			'Sales': row[3],
			'Revenue': row[4]
		}
	else:
		break

custom_printer.pprint(musicnotes_data)
print()
# -------------------------------------------------------------------------------------------------



# Load template file
xl_wb = load_workbook(template_xl_file)
sheet = xl_wb.worksheets[0]

# Define styles
boldunderline = Font(
	bold=True,
	underline='single'
)

grey_fill = PatternFill(
	patternType='solid',
	fgColor='808080'
)
yellow_fill = PatternFill(
	patternType='solid',
	fgColor='FFFF00'
)

thin_border = Side(border_style='thin')

# Write Sheet -----------------------------------------------------------------
starting_row = 4

# Write data
print('Writing data..')

for row, title in enumerate(master_list, starting_row):
	sheet[row][0].value = title
	title = title.lower()

	if title in musicnotes_data:
		title_data = musicnotes_data[title]
		sheet[row][1].value = title_data['Downloads']
		sheet[row][2].value = round(title_data['Sales'], 2)
		sheet[row][3].value = round(title_data['Revenue'], 2)
else:
	last_row_entry = row

for cell in sheet[last_row_entry+1]:
	cell.fill = grey_fill

# Write formula values for totals
totals_row = last_row_entry + 2

for cell in sheet[totals_row][1:]:
	col = cell.column_letter
	cell.value = f'=SUM({col}{starting_row}:{col}{last_row_entry})'

cell = sheet[totals_row][0]
cell.value = 'TOTAL:'
cell.font = boldunderline

# Write reporting period as header
header = sheet['B1']
header.value = quarter

if quarter == 'Q1':
	header.value = f'{year} {quarter}'
	header.fill = yellow_fill

# Style total revenue cell
total_revenue = sheet[totals_row][3]

total_revenue.fill = yellow_fill
total_revenue.border = Border(
	top=thin_border,
	right=thin_border,
	bottom=thin_border,
	left=thin_border
)

# Update sheet title
sheet.title = reporting_period
print()
# -------------------------------------------------------------------------------------------------

xl_wb.save(output_file)
